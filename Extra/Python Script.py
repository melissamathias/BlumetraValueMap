import openpyxl
import json

path = '/Users/krishiv/Test/relationship_chart_spreadsheet_final.xlsx'
file_handle = openpyxl.load_workbook(path)

sheet_obj = file_handle.active

data_list = []
data_list1 = []

header_var = '''
var width = 1600,
    height = 1600,
    box_width = 200,
    box_height = 30,
    box_heights = [40, 20, 40, 20],
    box_gap = {
        width: 150,
        height: 2
    },
    margin = {
        top: 100,
        right: 16,
        bottom: 16,
        left: 16
    },
    svg;

var data = {
    "nodes":'''

header_links =  ''',
    "links":'''
    
ending = '''
};
'''


final0 = '''
var nodes = [];
var links = [];
var lvlCount = 0;

//creates the horizontal link between two different objects
var link = d3.linkHorizontal()
    .x(function (d) {
        return d.y;
    })
    .y(function (d) {
        return d.x;
    });

//uses a for loop to find the specific node that is needed
function locate_node(text) {
    "use strict";
    var i;
    for (i = 0; i < nodes.length; i += 1) {
        if (nodes[i].name === text) {
            return nodes[i];
        }
    }
    return null;
}

//mouse action to allow the data to be dynamically changed
function mouse_action(val, stat, direction) {
    "use strict";
    d3.select("#" + val.id).classed("active", stat);

    links.forEach(function (d) {
        if (direction == "root") {
            if (d.source.id === val.id) {
                d3.select("#" + d.id).classed("activelink", stat);
                d3.select("#" + d.id).classed("link", !stat);
                if (d.target.lvl < val.lvl)
                    mouse_action(d.target, stat, "left");
                else if (d.target.lvl > val.lvl)
                    mouse_action(d.target, stat, "right");
            }
            if (d.target.id === val.id) {
                d3.select("#" + d.id).classed("activelink", stat);
                d3.select("#" + d.id).classed("link", !stat);
                if (direction == "root") {
                    if (d.source.lvl < val.lvl)
                        mouse_action(d.source, stat, "left");
                    else if (d.source.lvl > val.lvl)
                        mouse_action(d.source, stat, "right");
                }
            }
        } else if (direction == "left") {
            if (d.source.id === val.id && d.target.lvl < val.lvl) {
                d3.select("#" + d.id).classed("activelink", stat);
                d3.select("#" + d.id).classed("link", !stat);

                mouse_action(d.target, stat, direction);
            }
            if (d.target.id === val.id && d.source.lvl < val.lvl) {
                d3.select("#" + d.id).classed("activelink", stat);
                d3.select("#" + d.id).classed("link", !stat);
                mouse_action(d.source, stat, direction);
            }
        } else if (direction == "right") {
            if (d.source.id === val.id && d.target.lvl > val.lvl) {
                d3.select("#" + d.id).classed("activelink", stat);
                d3.select("#" + d.id).classed("link", !stat);
                mouse_action(d.target, stat, direction);
            }
            if (d.target.id === val.id && d.source.lvl > val.lvl) {
                d3.select("#" + d.id).classed("activelink", stat);
                d3.select("#" + d.id).classed("link", !stat);
                mouse_action(d.source, stat, direction);
            }
        }
    });
}

//when mouse clicks on it it stays as how it is
function mouse_action_click(val, stat, direction) {
    "use strict";
    d3.select("#" + val.id).classed("active_clk", stat);

    links.forEach(function (d) {
        if (direction == "root") {
            if (d.source.id === val.id) {
                d3.select("#" + d.id).classed("activelink_clk", stat);
                d3.select("#" + d.id).classed("link_clk", !stat);
                if (d.target.lvl < val.lvl)
                    mouse_action_click(d.target, stat, "left");
                else if (d.target.lvl > val.lvl)
                    mouse_action_click(d.target, stat, "right");
            }
            if (d.target.id === val.id) {
                d3.select("#" + d.id).classed("activelink_clk", stat);
                d3.select("#" + d.id).classed("link_clk", !stat);
                if (direction == "root") {
                    if (d.source.lvl < val.lvl)
                        mouse_action_click(d.source, stat, "left");
                    else if (d.source.lvl > val.lvl)
                        mouse_action_click(d.source, stat, "right");
                }
            }
        } else if (direction == "left") {
            if (d.source.id === val.id && d.target.lvl < val.lvl) {
                d3.select("#" + d.id).classed("activelink_clk", stat);
                d3.select("#" + d.id).classed("link_clk", !stat);

                mouse_action_click(d.target, stat, direction);
            }
            if (d.target.id === val.id && d.source.lvl < val.lvl) {
                d3.select("#" + d.id).classed("activelink_clk", stat);
                d3.select("#" + d.id).classed("link_clk", !stat);
                mouse_action_click(d.source, stat, direction);
            }
        } else if (direction == "right") {
            if (d.source.id === val.id && d.target.lvl > val.lvl) {
                d3.select("#" + d.id).classed("activelink_clk", stat);
                d3.select("#" + d.id).classed("link_clk", !stat);
                mouse_action_click(d.target, stat, direction);
            }
            if (d.target.id === val.id && d.source.lvl > val.lvl) {
                d3.select("#" + d.id).classed("activelink_clk", stat);
                d3.select("#" + d.id).classed("link_clk", !stat);
                mouse_action_click(d.source, stat, direction);
            }
        }
    });
}

function untravel_links() {
    "use strict";
    links.forEach(function (d) {
        d.visited = false;
    });
}

function clearActiveLinksNodes() {
    d3.selectAll(".active_clk").classed("active_clk", false);
    d3.selectAll(".activelink_clk").classed("activelink_clk", false);

}

function create_relationship_diagram(data) {
    "use strict";
    var count = [];
    var col_lengths = [];

    data.nodes.forEach(function (d) {
        count[d.lvl] = 0;
        col_lengths[d.lvl] = 0;
    });
    lvlCount = count.length;

    //offsetShortColumns
    data.nodes.forEach(function (d, i) {
        col_lengths[d.lvl] += 1;
    });

    let max_col_length = Math.max(...col_lengths);
    let max_col_lvl = col_lengths.indexOf(max_col_length);

    let tallest_col_height = max_col_length * (box_heights[max_col_lvl] + box_gap.height);

    data.nodes.forEach(function (d, i) {
        d.x = margin.left + d.lvl * (box_width + box_gap.width);
        d.y = (margin.top + (box_heights[d.lvl] + box_gap.height) * count[d.lvl]);

        if (d.lvl == 0 ) {
            d.y += ((tallest_col_height - (col_lengths[d.lvl]) * (box_heights[d.lvl] + box_gap.height))/2);
            }
        if (d.lvl == 2) {
            d.y += ((tallest_col_height - (col_lengths[d.lvl]) * (box_heights[d.lvl] + box_gap.height))/2);        
        }
        if(d.lvl == 1){
        d.y += ((tallest_col_height - (col_lengths[d.lvl]) * (box_heights[d.lvl] + box_gap.height))/2);
        }

        d.id = "n" + i;
        count[d.lvl] += 1;
        nodes.push(d);
    });

    data.links.forEach(function (d) {
        links.push({
            source: locate_node(d.source),
            target: locate_node(d.target),
            id: "l" + locate_node(d.source).id + locate_node(d.target).id
        });
    });
    untravel_links();

    d3.select("#chart_canvas")
        .on("click", clearActiveLinksNodes)

    svg.append("g")
        .attr("class", "nodes");

    var node = svg.select(".nodes")
        .selectAll("g")
        .data(nodes)
        .enter()
        .append("g")
        .attr("class", "unit");

    node.append("rect")
        .attr("x", function (d) { return d.x; })
        .attr("y", function (d) { return d.y; })
        .attr("id", function (d) { return d.id; })
        .attr("width", box_width)
        .attr("height", function (d) { return box_heights[d.lvl]; })
        .attr("class", "node")
        .attr("class", function (d) { return "col" + d.lvl; })
        .attr("rx", 0.03*box_width)
        .attr("ry", 0.2*box_height)
        .on("mouseover", function () {
            mouse_action(d3.select(this).datum(), true, "root");
            untravel_links();
        })
        .on("mouseout", function () {
            mouse_action(d3.select(this).datum(), false, "root");
            untravel_links();
        })
        .on("click", function (e) {
            e.stopPropagation();
            if (d3.select(this).classed("active_clk")) {
                mouse_action_click(d3.select(this).datum(), false, "root");
            }
            else {
                mouse_action_click(d3.select(this).datum(), true, "root");
            }
            untravel_links();
        });
      
        
        
'''


final1 = '''



var businessValueNodesCount = nodes.filter(function(d) {
    return d.lvl == 0; // Adjust 0 to the appropriate level number for "Business Value"
}).length;

node.each(function(d, i) {
    if (d.lvl == 1) {
        var smallerBox = d3.select(this)
            .append("rect")
            .attr("x", d.x + 0.95 * box_width) // Position it to the right side
            .attr("y", d.y)
            .attr("id", d.id + "-smaller-box")
            .attr("width", 0.05*box_width)
            .attr("height", 0.66*box_height)
            .attr("class", "smaller-box")
            .attr("rx", 2*box_width)
            .attr("ry", 0.66*box_height)
            
            
       d3.select(this)
    .append("text")
    .attr("x", d.x + 0.972 * box_width) // Adjust position for text
    .attr("y", d.y + 0.25 * box_height) // Adjust position for text
    .attr("text-anchor", "middle") // Center text horizontally
    .attr("alignment-baseline", "middle") // Center text vertically
    .text("...")
    .style("cursor", "pointer") // Set cursor to pointer
    .on("click", function(d) {
        window.open(boxData[i-businessValueNodesCount].link, "_blank");
    });

            
            
    }
});

    node.append("text")
        .attr("class", function (d) { return "label" + d.lvl; })
        .attr("x", function (d) { return d.x + (0.055)*box_width; })
        .attr("y", function (d) { return d.y + (0.5)*box_height; })
        .text(function (d) { return d.name; });



    //overall title and column titles
    svg.append("text")
        .attr("id", "chart_title")
        .attr("x", 0.00625*width)
        .attr("y", 0.015625*height)
        .text("Customer(Org) Data - Value Map");
        
        
        


    var column_data = [
        {
            "lvl": 0,
            "title": "Business Value"
        },
        {
            "lvl": 1,
            "title": "Use Cases"
        },
        {
            "lvl": 2,
            "title": "Capabilities"
        },
        {
            "lvl": 3,
            "title": "Features"
        },


    ]
    var col_titles = svg.append("g")
        .attr("class", "column_title")
        .selectAll("g")
        .data(column_data)
        .enter()
        .append("g")

    col_titles.append("rect")
        .attr("x", function (d) { return margin.left + d.lvl * (box_width + box_gap.width); })
        .attr("y", function (d) { return 0.0375*height; })
        .attr("id", function (d) { return "col_title" + d.lvl; })
        .attr("width", box_width)
        .attr("height", function (d) { return box_height; })
        .attr("class", function (d) { return "col_title"; })
        /* .attr("rx", 2*box_height)
        .attr("ry", 0.03*box_width) */

    col_titles.append("text")
        .attr("x", function (d) { return 0.83333333333*box_height + margin.left + d.lvl * (box_width + box_gap.width); })
        .attr("y", function (d) { return 0.0375*height + 0.66*box_height; })
        .attr("class", "col_title_text")
        .text(function (d) { return d.title; });



    //d.x = margin.left + d.lvl * (box_width + box_gap.width);
    // d.y = (margin.top + (box_heights[d.lvl] + box_gap.height) * count[d.lvl]);



    links.forEach(function (li) {
        svg.append("path", "g")
            .attr("class", "link")
            .attr("id", li.id)
            .attr("d", function () {
                var target_link_point = {
                    //x: li.target.y + 0.5 * box_height,
                    x: li.target.y + 0.5 * box_heights[li.target.lvl],
                    y: li.target.x
                };
                var source_link_point = {
                    x: li.source.y + 0.5 * box_heights[li.source.lvl],
                    y: li.source.x
                };

                if (source_link_point.y < target_link_point.y) {
                    source_link_point.y += box_width;
                } else {
                    target_link_point.y += box_width;
                }
                return link({
                    source: source_link_point,
                    target: target_link_point
                });
            });
    });
}

svg = d3.select("#rel_diagram").append("svg")
    .attr("width", width)
    .attr("height", height)
    .attr("id", "chart_canvas")
    .append("g");

svg.append("image")
   .attr("xlink:href", "blumetra-logo.svg")
   .attr("x", 0.774*width)
   .attr("y", 0.0000000000000000000000000001*height)
   .attr("width", 200) 
   .attr("height", 36);


create_relationship_diagram(data);
'''



workbook = openpyxl.load_workbook('/Users/krishiv/Test/relationship_chart_spreadsheet_final.xlsx')
sheet_name = 'Use Cases Nodes'
sheet = workbook[sheet_name]
column2_data = []
for row in sheet.iter_rows(min_row=2, min_col=2, max_col=2, values_only=True):
    column2_data.append(row[0])


boxdata1 = "var boxData = "
links_data = [{'link': link} for link in column2_data]
boxdata3 = "; "


json_string = json.dumps(links_data)



# Set to keep track of already added cell values
added_values = set()
added_source_target = set()
nodes0_list = []
nodes1_list = []
nodes2_list = []
nodes3_list = []

column_print = 2

for sheet_name in file_handle.sheetnames:
    sheet_obj = file_handle[sheet_name]
    
    if sheet_obj == file_handle["lvl2 to lvl3"]:
        column_print = sheet_obj.max_column + 1
    for col in range(1, column_print):
        lvl_value = (sheet_obj.cell(row=1, column=col).value)
        for row in range(2, sheet_obj.max_row + 1):
            cell_value = sheet_obj.cell(row=row, column=col).value
            if cell_value is not None and cell_value not in added_values:
                    data_dict = {
                        "lvl": lvl_value,  
                        "name": cell_value
                    }
                    data_list.append(data_dict)
                    added_values.add(cell_value)

        column_print1 = sheet_obj.max_column 
        if sheet_obj == file_handle["Use Cases Nodes"]:
            continue
        for col in range(1, column_print1):
            for row in range(2, sheet_obj.max_row + 1):
                source_value = sheet_obj.cell(row=row, column=col).value
                target_value = sheet_obj.cell(row=row, column=col+1).value
                if source_value is not None and target_value is not None:
                    pair = (source_value, target_value)
                    if pair not in added_source_target:
                        data_dict1 = {
                            "source": source_value,  
                            "target": target_value
                        }
                        data_list1.append(data_dict1)
                        added_source_target.add(pair)
                        

json_data = json.dumps(data_list, indent=4)
json_data1 = json.dumps(data_list1, indent=4)

output_path = '/Users/krishiv/Test/d3_relationship_chartv6.js'  
    
with open(output_path, 'w') as output_file:
    output_file.write(header_var)
    output_file.write(json_data)
    output_file.write(header_links)
    output_file.write(json_data1)
    output_file.write(ending)
    output_file.write(final0)
    output_file.write(boxdata1)
    output_file.write(json_string)
    output_file.write(boxdata3)
    output_file.write(final1)

print(f'Data has been written to {output_path}')
